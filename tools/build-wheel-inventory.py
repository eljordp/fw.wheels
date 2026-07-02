#!/usr/bin/env python3
"""Build FW Wheels storefront inventory data from vendor spreadsheet exports.

Usage:
  python3 tools/build-wheel-inventory.py
  FW_INVENTORY_SOURCE_DIR="/path/to/fw-wheels-drive-audit" python3 tools/build-wheel-inventory.py

The script reads the three wheel sources JP opened in Drive:
- Fw.wheelz@gmail.com Inventory / MRR-Vors export
- A Spec Wheels Inventory / AodHan + MRR export
- MFlow Racing Dealer copied TSV tabs

It writes:
- data/wheel-inventory.json
- data/wheel-inventory-report.md
"""

from __future__ import annotations

import csv
import json
import os
import re
import urllib.request
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_DIR = Path("/Users/jp/Documents/New project/tmp/fw-wheels-drive-audit")
SOURCE_DIR = Path(os.environ.get("FW_INVENTORY_SOURCE_DIR", DEFAULT_SOURCE_DIR))

SOURCE_EXPORTS = SOURCE_DIR / "01-source-exports"
EXTRACTED_TABS = SOURCE_DIR / "02-extracted-tabs"
OUT_JSON = ROOT / "data" / "wheel-inventory.json"
OUT_REPORT = ROOT / "data" / "wheel-inventory-report.md"
OFFICIAL_CATALOG_QTY = 20

MODEL_TO_SLUG = {
    ("AODHAN", "AH01"): "ah01",
    ("AODHAN", "AH02"): "ah02",
    ("AODHAN", "AH03"): "ah03",
    ("AODHAN", "AH04"): "ah04",
    ("AODHAN", "AH05"): "ah05",
    ("AODHAN", "AH06"): "ah06",
    ("AODHAN", "AH07"): "ah07",
    ("AODHAN", "AH08"): "ah08",
    ("AODHAN", "AH09"): "ah09",
    ("AODHAN", "AHX"): "ahx",
    ("AODHAN", "AH11"): "ah11",
    ("AODHAN", "DS01"): "ds01",
    ("AODHAN", "DS02"): "ds02",
    ("AODHAN", "DS03"): "ds03",
    ("AODHAN", "DS05"): "ds05",
    ("AODHAN", "DS06"): "ds06",
    ("AODHAN", "DS07"): "ds07",
    ("AODHAN", "DS08"): "ds08",
    ("AODHAN", "DS09"): "ds09",
    ("AODHAN", "DSX"): "dsx",
    ("AODHAN", "AFF1"): "aff1",
    ("AODHAN", "AFF2"): "aff2",
    ("AODHAN", "AFF3"): "aff3",
    ("AODHAN", "AFF7"): "aff7",
    ("AODHAN", "AFF9"): "aff9",
    ("MFLOW", "MFR1"): "mfr1",
    ("MFLOW", "MFR2"): "mfr2",
    ("MFLOW", "MFR3"): "mfr3",
    ("MFLOW", "MFR4"): "mfr4",
    ("MFLOW", "MFL1"): "mfl1",
    ("MFLOW", "MFL2"): "mfl2",
    ("MFLOW", "MF01"): "mf01",
    ("MFLOW", "MF02"): "mf02",
    ("MFLOW", "MF03"): "mf03",
    ("MFLOW", "MF04"): "mf04",
    ("MFLOW", "MF05"): "mf05",
    ("MFLOW", "MF06"): "mf06",
    ("VORS", "TR4"): "vors-tr4",
    ("VORS", "TR10"): "vors-tr10",
    ("VORS", "TR14"): "vors-tr14",
    ("VORS", "TR37"): "vors-tr37",
    ("VORS", "TR88"): "vors-tr88",
    ("VORS", "VR8"): "vors-vr8",
    ("VORS", "AR1"): "vors-ar1",
    ("VORS", "AR5"): "vors-ar5",
    ("VORS", "SP1"): "vors-sp1",
    ("VORS", "LT53"): "vors-lt53",
    ("VORS", "UO2"): "vors-uo2",
}

# Owner/business rules win over broad distributor catalog rows when needed.
OVERRIDES = {}

OFFICIAL_CATALOG_PRODUCTS = [
    ("AODHAN", "AH09", "ah09", "https://www.aodhanwheels.com/products/aodhan-wheels-ah09.js"),
    ("AODHAN", "AH11", "ah11", "https://www.aodhanwheels.com/products/aodhan-wheels-ah11.js"),
    ("AODHAN", "AHX", "ahx", "https://www.aodhanwheels.com/products/aodhan-wheels-ahx.js"),
    ("AODHAN", "DS03", "ds03", "https://www.aodhanwheels.com/products/aodhan-wheels-ds03.js"),
    ("AODHAN", "DSX", "dsx", "https://www.aodhanwheels.com/products/aodhan-wheels-dsx.js"),
    ("VORS", "UO2", "vors-uo2", "https://www.vorswheels.com/products/uo2.js"),
]


@dataclass
class SkuRow:
    brand: str
    model: str
    slug: str
    size: str
    finish: str
    sku: str = ""
    bolt_pattern: str = ""
    offset: str = ""
    center_bore: str = ""
    price: float | None = None
    qty: float = 0
    image: str = ""
    source: str = ""
    source_status: str = "sourced"


@dataclass
class VariantRollup:
    finishes: set[str] = field(default_factory=set)
    available_finishes: set[str] = field(default_factory=set)
    bolt_patterns: set[str] = field(default_factory=set)
    offsets: set[str] = field(default_factory=set)
    bolt_configs: list[dict[str, str]] = field(default_factory=list)
    price_configs: list[dict[str, Any]] = field(default_factory=list)
    finish_prices: dict[str, float] = field(default_factory=dict)
    price: float | None = None
    qty: float = 0
    skus: list[str] = field(default_factory=list)
    image: str = ""


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_number(value: Any) -> float:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def normalize_finish(value: str) -> str:
    text = clean(value)
    text = re.sub(r"\bMatt\b", "Matte", text, flags=re.I)
    text = re.sub(r"\bMachine\b", "Machined", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    replacements = {
        "HYPER SILVER MACHINED FACE": "Hyper Silver Machined Face",
        "HYPER SILVER MACHINED TIP": "Hyper Silver Machined Tip",
        "MATTE BLACK MACHINED TIP": "Matte Black Machined Tip",
        "MATTE BRONZE MACHINED TIP": "Matte Bronze Machined Tip",
        "MACHINED SILVER": "Machined Silver",
        "MATTE BLACK": "Matte Black",
        "MATTE BRONZE": "Matte Bronze",
    }
    return replacements.get(text.upper(), text)


def normalize_offset(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    try:
        n = int(float(text))
        return f"+{n}" if n > 0 else str(n)
    except ValueError:
        return text if text.startswith(("+", "-")) else f"+{text}" if text.isdigit() else text


def normalize_size(value: Any) -> str:
    return re.sub(r"\s*\([^)]*\)", "", clean(value)).strip()


def site_slug(brand: str, model: str) -> str | None:
    return MODEL_TO_SLUG.get((brand.upper(), model.upper()))


def split_dual_bolt(value: str) -> list[str]:
    text = clean(value).replace(" ", "")
    if not text:
        return []
    if "/" not in text:
        return [text]
    prefix = text.split("x", 1)[0] + "x" if "x" in text else ""
    head, tail = text.split("/", 1)
    if "x" not in tail and prefix:
        return [head, prefix + tail]
    return [head, tail]


def add_bolt_configs(variant: VariantRollup, bolt_pattern: str, offset: str, center_bore: str) -> None:
    bolts = split_dual_bolt(bolt_pattern)
    if not bolts and bolt_pattern:
        bolts = [bolt_pattern]
    for bolt in bolts:
        item = {"bolt": bolt, "offset": offset}
        if center_bore:
            item["cb"] = center_bore.replace("mm", "")
        if item not in variant.bolt_configs:
            variant.bolt_configs.append(item)


def add_price_configs(variant: VariantRollup, finish: str, bolt_pattern: str, offset: str, price: float | None) -> None:
    if not price:
        return
    if not variant.price or price < variant.price:
        variant.price = price
    if finish and (finish not in variant.finish_prices or price < variant.finish_prices[finish]):
        variant.finish_prices[finish] = price

    bolts = split_dual_bolt(bolt_pattern)
    if not bolts and bolt_pattern:
        bolts = [bolt_pattern]
    if not bolts:
        bolts = [""]
    for bolt in bolts:
        item = {
            "finish": finish,
            "bolt": bolt,
            "offset": offset,
            "price": price,
        }
        if item not in variant.price_configs:
            variant.price_configs.append(item)


def header_index(headers: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, header in enumerate(headers):
        if header and header not in idx:
            idx[header] = i
    return idx


def load_workbook_rows(path: Path, sheet_name: str, required_headers: tuple[str, ...] = ()) -> tuple[list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        headers = [clean(v) for v in row]
        if all(header in headers for header in required_headers):
            return headers, rows[i + 1:]
    headers = [clean(v) for v in rows[0]]
    return headers, rows[1:]


def aodhan_rows() -> list[SkuRow]:
    path = SOURCE_EXPORTS / "A Spec Wheels Inventory.xlsx"
    headers, rows = load_workbook_rows(path, "AodHan", ("Part #", "Model", "Size", "Color"))
    idx = header_index(headers)
    out: list[SkuRow] = []
    for row in rows:
        model = clean(row[idx.get("Model", -1)]).upper()
        slug = site_slug("AODHAN", model)
        if not slug:
            continue
        finish = normalize_finish(clean(row[idx.get("Color", -1)]))
        size = clean(row[idx.get("Size", -1)])
        if not finish or not size:
            continue
        qty = parse_number(row[idx.get("On Hand", -1)])
        qty += parse_number(row[idx.get("Incoming\n(1 - 2 Weeks)", -1)])
        qty += parse_number(row[idx.get("Incoming\n(3 - 5 Weeks)", -1)])
        out.append(SkuRow(
            brand="AODHAN",
            model=model,
            slug=slug,
            size=size,
            finish=finish,
            sku=clean(row[idx.get("Part #", -1)]),
            bolt_pattern=clean(row[idx.get("PCD", -1)]),
            offset=normalize_offset(row[idx.get("Offset", -1)]),
            center_bore=clean(row[idx.get("Hub Bore", -1)]),
            price=parse_number(row[idx.get("MAP (Resale)", -1)]) or parse_number(row[idx.get("MSRP", -1)]) or None,
            qty=qty,
            image=clean(row[idx.get("Picture", -1)]),
            source="A Spec Wheels Inventory.xlsx / AodHan",
        ))
    return out


def vors_rows() -> list[SkuRow]:
    sources = [
        (SOURCE_EXPORTS / "Fw.wheelz@gmail.com Inventory.xlsx", "Sheet1", "Fw.wheelz@gmail.com Inventory.xlsx / Sheet1"),
        (SOURCE_EXPORTS / "A Spec Wheels Inventory.xlsx", "MRR", "A Spec Wheels Inventory.xlsx / MRR"),
    ]
    out: list[SkuRow] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for path, sheet, label in sources:
        headers, rows = load_workbook_rows(path, sheet, ("Brand", "Model", "Size"))
        idx = header_index(headers)
        for row in rows:
            brand = clean(row[idx.get("Brand", -1)]).upper()
            model = clean(row[idx.get("Model", -1)]).upper()
            slug = site_slug(brand, model)
            if brand != "VORS" or not slug:
                continue
            finish = normalize_finish(clean(row[idx.get("Finish Long Description", -1)]) or clean(row[idx.get("Finish", -1)]))
            size = clean(row[idx.get("Size", -1)])
            sku = clean(row[idx.get("SKU", -1)])
            qty = parse_number(row[idx.get("Inventory", -1)] if "Inventory" in idx else row[idx.get("Qty", -1)])
            key = (slug, size, finish, sku, clean(row[idx.get("PCD", -1)]), normalize_offset(row[idx.get("ET", -1)]))
            if key in seen:
                continue
            seen.add(key)
            out.append(SkuRow(
                brand="VORS",
                model=model,
                slug=slug,
                size=size,
                finish=finish,
                sku=sku,
                bolt_pattern=clean(row[idx.get("PCD", -1)]),
                offset=normalize_offset(row[idx.get("ET", -1)]),
                center_bore=clean(row[idx.get("CB", -1)]),
                price=parse_number(row[idx.get("MAP", -1)]) or parse_number(row[idx.get("MSRP", -1)]) or None,
                qty=qty,
                image=clean(row[idx.get("Pic Link 1", -1)]),
                source=label,
            ))
    return out


def mflow_rows_from_tsv(path: Path, source: str) -> list[SkuRow]:
    rows = list(csv.reader(path.read_text(errors="replace").splitlines(), delimiter="\t"))
    header_i = None
    for i, row in enumerate(rows[:25]):
        lowered = [clean(cell).lower() for cell in row]
        if "part #" in lowered and "description" in lowered and "in stock" in lowered:
            header_i = i
            break
    if header_i is None:
        return []

    headers = [clean(v) for v in rows[header_i]]
    idx = {h: i for i, h in enumerate(headers)}
    out: list[SkuRow] = []
    for row in rows[header_i + 1:]:
        row += [""] * (len(headers) - len(row))
        description = clean(row[idx.get("Description", -1)])
        model = parse_mflow_model(description)
        slug = site_slug("MFLOW", model)
        if not slug:
            continue
        finish = parse_mflow_finish(model, description)
        if not finish:
            continue
        out.append(SkuRow(
            brand="MFLOW",
            model=model,
            slug=slug,
            size=parse_mflow_size(description),
            finish=normalize_finish(finish),
            sku=clean(row[idx.get("Part #", -1)]),
            bolt_pattern=parse_mflow_bolt(description),
            offset=parse_mflow_offset(description),
            center_bore=parse_mflow_cb(description),
            price=parse_number(row[idx.get("MAP(retail price)", -1)]) or parse_number(row[idx.get("MSRP", -1)]) or None,
            qty=max(0, parse_number(row[idx.get("In Stock", -1)])),
            image="",
            source=source,
        ))
    return out


def parse_mflow_size(description: str) -> str:
    match = re.search(r"\b(\d{2})X(\d{1,2}(?:\.\d)?)\b", description, re.I)
    return f"{match.group(1)}x{match.group(2)}" if match else ""


def parse_mflow_model(description: str) -> str:
    match = re.search(r"\bMFLOW\s+([A-Z]{2,3}\d{1,2})\b", description, re.I)
    return match.group(1).upper() if match else ""


def parse_mflow_bolt(description: str) -> str:
    match = re.search(r"\b(\dX\d{3}(?:\.\d)?)\b", description, re.I)
    return match.group(1).lower().replace("x", "x") if match else ""


def parse_mflow_offset(description: str) -> str:
    match = re.search(r"\s([+-]\d{1,3})(?:\s*CB|\s+\d{2,3}(?:\.\d)?)", description, re.I)
    return normalize_offset(match.group(1)) if match else ""


def parse_mflow_cb(description: str) -> str:
    match = re.search(r"\bCB\s*(\d{2,3}(?:\.\d)?)", description, re.I)
    if match:
        return match.group(1)
    match = re.search(r"\s[+-]\d{1,3}\s+(\d{2,3}(?:\.\d)?)\s+", description, re.I)
    return match.group(1) if match else ""


def parse_mflow_finish(model: str, description: str) -> str:
    text = description.upper()
    if "HYPER SILVER MACHINED FACE" in text:
        return "Hyper Silver Machined Face"
    if "HYPER SILVER MACHINED TIP" in text:
        return "Hyper Silver Machined Tip"
    if "MATTE BLACK MACHINED TIP" in text:
        return "Matte Black Machined Tip"
    if "MATTE BRONZE MACHINED TIP" in text:
        return "Matte Bronze Machined Tip"
    if "CHROME (PVD)" in text:
        return "Chrome (PVD)"
    if "MATT BLACK MACHINED LIP" in text or "MATTE BLACK MACHINED LIP" in text:
        return "Matte Black Machined Lip"
    if "MATT BRONZE MACHINED LIP" in text or "MATTE BRONZE MACHINED LIP" in text:
        return "Matte Bronze Machined Lip"
    if "MACHINED SILVER" in text:
        return "Machined Silver"
    if "HYPER BLACK" in text:
        return "Hyper Black"
    if "HYPER SILVER" in text:
        return "Hyper Silver"
    if "GLOSS BLACK" in text:
        return "Gloss Black"
    if "MATT BLACK" in text or "MATTE BLACK" in text:
        return "Matte Black"
    if "MATT BRONZE" in text or "MATTE BRONZE" in text:
        return "Matte Bronze"
    return ""


def mflow_rows() -> list[SkuRow]:
    return (
        mflow_rows_from_tsv(EXTRACTED_TABS / "MFlow Racing Inventory.tsv", "MFlow Racing Dealer / MFlow Racing Inventory")
        + mflow_rows_from_tsv(EXTRACTED_TABS / "MFlow OFF-ROAD Inventory.tsv", "MFlow Racing Dealer / MFlow OFF-ROAD Inventory")
    )


def official_catalog_rows() -> list[SkuRow]:
    out: list[SkuRow] = []
    for brand, model, slug, url in OFFICIAL_CATALOG_PRODUCTS:
        with urllib.request.urlopen(url, timeout=30) as response:
            product = json.load(response)
        option_positions = {opt["name"].lower(): opt["position"] for opt in product.get("options", [])}
        size_pos = option_positions.get("size")
        color_pos = option_positions.get("color")
        pcd_pos = next((pos for name, pos in option_positions.items() if "pcd" in name), None)
        image = clean(product.get("featured_image"))

        for variant in product.get("variants", []):
            if not variant.get("available"):
                continue
            size = normalize_size(variant.get(f"option{size_pos}")) if size_pos else ""
            finish = normalize_finish(clean(variant.get(f"option{color_pos}"))) if color_pos else ""
            pcd = clean(variant.get(f"option{pcd_pos}")) if pcd_pos else ""
            match = re.match(r"(\S+)\s*(?:\|\s*)?([+-]?\d+(?:\.\d+)?)\s*(?:\|\s*)?(\d{2,3}(?:\.\d+)?)?", pcd)
            bolt_pattern = match.group(1) if match else pcd
            offset = normalize_offset(match.group(2)) if match else ""
            center_bore = match.group(3) if match and match.lastindex and match.lastindex >= 3 else ""
            variant_image = variant.get("featured_image") or {}
            row_image = clean(variant_image.get("src") if isinstance(variant_image, dict) else "") or image
            if not size or not finish:
                continue
            out.append(SkuRow(
                brand=brand,
                model=model,
                slug=slug,
                size=size,
                finish=finish,
                sku=clean(variant.get("sku")),
                bolt_pattern=bolt_pattern,
                offset=offset,
                center_bore=center_bore,
                price=parse_number(variant.get("price")) / 100 if variant.get("price") else None,
                qty=OFFICIAL_CATALOG_QTY,
                image=row_image,
                source=f"Official brand catalog / {url.replace('.js', '')}",
                source_status="official_brand_catalog",
            ))
    return out


def apply_overrides(rows: list[SkuRow]) -> list[SkuRow]:
    out = []
    for row in rows:
        override = OVERRIDES.get(row.slug)
        if override:
            allowed = {normalize_finish(f).lower() for f in override["allowed_finishes"]}
            if normalize_finish(row.finish).lower() not in allowed:
                continue
        out.append(row)
    return out


def build_inventory(rows: list[SkuRow]) -> dict[str, Any]:
    products: dict[str, dict[str, Any]] = {}
    rollups: dict[str, dict[str, VariantRollup]] = defaultdict(lambda: defaultdict(VariantRollup))
    product_sources: dict[str, set[str]] = defaultdict(set)

    for row in rows:
        variant = rollups[row.slug][row.size]
        variant.finishes.add(row.finish)
        if row.qty > 0:
            variant.available_finishes.add(row.finish)
        if row.bolt_pattern:
            variant.bolt_patterns.add(row.bolt_pattern)
        if row.offset:
            variant.offsets.add(row.offset)
        add_bolt_configs(variant, row.bolt_pattern, row.offset, row.center_bore)
        variant.qty += row.qty
        if row.sku and row.sku not in variant.skus:
            variant.skus.append(row.sku)
        add_price_configs(variant, row.finish, row.bolt_pattern, row.offset, row.price)
        if row.image and not variant.image:
            variant.image = row.image
        product_sources[row.slug].add(row.source)
        products.setdefault(row.slug, {
            "brand": row.brand,
            "model": row.model,
            "slug": row.slug,
            "sourceStatus": row.source_status,
        })

    for slug, product in products.items():
        product["sources"] = sorted(product_sources[slug])
        product["variants"] = {}
        for size, variant in sorted(rollups[slug].items()):
            catalog_finishes = sorted(variant.finishes)
            available_finishes = sorted(variant.available_finishes)
            qty = OFFICIAL_CATALOG_QTY if product["sourceStatus"] == "official_brand_catalog" else variant.qty
            product["variants"][size] = {
                "size": size,
                "finishes": available_finishes if available_finishes else catalog_finishes,
                "catalogFinishes": catalog_finishes,
                "availableFinishes": available_finishes,
                "boltPatterns": sorted(variant.bolt_patterns),
                "offsets": sorted(variant.offsets, key=lambda v: parse_number(v)),
                "boltConfigs": variant.bolt_configs,
                "priceConfigs": variant.price_configs,
                "finishPrices": dict(sorted(variant.finish_prices.items())),
                "price": variant.price,
                "qty": int(qty) if qty == int(qty) else qty,
                "stockStatus": "in_stock" if qty > 0 else "sold_out",
                "skus": variant.skus[:30],
                "image": variant.image,
            }
        override = OVERRIDES.get(slug)
        if override:
            product["override"] = override

    for (_brand, _model), slug in MODEL_TO_SLUG.items():
        if slug not in products:
            products[slug] = {
                "brand": _brand,
                "model": _model,
                "slug": slug,
                "sourceStatus": "no_source_rows",
                "sources": [],
                "variants": {},
            }

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceDir": str(SOURCE_DIR),
        "policy": {
            "hideNoSourceModels": True,
            "hideUnsupportedSizes": True,
            "useAvailableFinishesWhenPresent": True,
            "officialCatalogQty": OFFICIAL_CATALOG_QTY,
            "ownerOverrides": OVERRIDES,
        },
        "products": dict(sorted(products.items())),
    }


def write_report(payload: dict[str, Any]) -> None:
    products = payload["products"]
    sourced = [p for p in products.values() if p["sourceStatus"] == "sourced"]
    official = [p for p in products.values() if p["sourceStatus"] == "official_brand_catalog"]
    missing = [p for p in products.values() if p["sourceStatus"] == "no_source_rows"]
    lines = [
        "# FW Wheels Spreadsheet Inventory Build",
        "",
        f"Generated: `{payload['generatedAt']}`",
        f"Source folder: `{payload['sourceDir']}`",
        "",
        "## Summary",
        f"- Site models tracked: {len(products)}",
        f"- Models with spreadsheet rows: {len(sourced)}",
        f"- Models added from official brand catalog: {len(official)}",
        f"- Official catalog temporary quantity: {OFFICIAL_CATALOG_QTY}",
        f"- Models hidden because no source rows: {len(missing)}",
        "",
        "## Hidden Because No Source Rows",
    ]
    if missing:
        for product in missing:
            lines.append(f"- {product['brand']} {product['model']} (`{product['slug']}`)")
    else:
        lines.append("- None")
    lines += ["", "## Owner Overrides"]
    for slug, override in OVERRIDES.items():
        lines.append(f"- `{slug}`: allowed finishes = {', '.join(override['allowed_finishes'])}; reason = {override['reason']}")
    lines += ["", "## Sourced Models"]
    for product in sourced:
        variant_count = len(product["variants"])
        finish_count = len({f for variant in product["variants"].values() for f in variant["finishes"]})
        lines.append(f"- {product['brand']} {product['model']} (`{product['slug']}`): {variant_count} sizes, {finish_count} sellable finishes")
    lines += ["", "## Official Brand Catalog Models"]
    if official:
        for product in official:
            variant_count = len(product["variants"])
            finish_count = len({f for variant in product["variants"].values() for f in variant["finishes"]})
            lines.append(f"- {product['brand']} {product['model']} (`{product['slug']}`): {variant_count} sizes, {finish_count} sellable finishes")
    else:
        lines.append("- None")
    OUT_REPORT.write_text("\n".join(lines) + "\n")


def main() -> None:
    required = [
        SOURCE_EXPORTS / "A Spec Wheels Inventory.xlsx",
        SOURCE_EXPORTS / "Fw.wheelz@gmail.com Inventory.xlsx",
        EXTRACTED_TABS / "MFlow Racing Inventory.tsv",
        EXTRACTED_TABS / "MFlow OFF-ROAD Inventory.tsv",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing source export(s):\n" + "\n".join(missing))

    rows = apply_overrides(aodhan_rows() + vors_rows() + mflow_rows() + official_catalog_rows())
    payload = build_inventory(rows)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, indent=2) + "\n")
    write_report(payload)

    sourced = sum(1 for p in payload["products"].values() if p["sourceStatus"] == "sourced")
    official = sum(1 for p in payload["products"].values() if p["sourceStatus"] == "official_brand_catalog")
    hidden = sum(1 for p in payload["products"].values() if p["sourceStatus"] == "no_source_rows")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_REPORT}")
    print(f"Sourced models: {sourced}; official catalog models: {official}; hidden no-source models: {hidden}")


if __name__ == "__main__":
    main()
